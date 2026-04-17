import pandas as pd
import re
import sys
import io
from difflib import SequenceMatcher

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# === 1. Load data ===
dj = pd.read_excel('C:/Projects/legend-team/Data/20260402_데일리잡 리스트_검토완료.xlsx', dtype=str)
hsm_grp = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='그룹사', dtype=str)
hsm_sub = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='계열사', dtype=str)

print(f"데일리잡: {len(dj)}건")
print(f"HSM 그룹사: {len(hsm_grp)}건 (고유 사업자번호: {hsm_grp['사업자번호'].nunique()})")
print(f"HSM 계열사: {len(hsm_sub)}건 (고유 사업자번호: {hsm_sub['사업자번호'].nunique()})")

# === 2. Normalize 사업자번호 ===
def normalize_bizno(val):
    if pd.isna(val):
        return None
    return re.sub(r'[^0-9]', '', str(val).strip())

dj['사업자번호_norm'] = dj['사업자번호'].apply(normalize_bizno)
hsm_grp['사업자번호_norm'] = hsm_grp['사업자번호'].apply(normalize_bizno)
hsm_sub['사업자번호_norm'] = hsm_sub['사업자번호'].apply(normalize_bizno)

# === 3. Deduplicate HSM by 사업자번호 (take first row per company) ===
hsm_grp_dedup = hsm_grp.drop_duplicates(subset='사업자번호_norm', keep='first').copy()
hsm_sub_dedup = hsm_sub.drop_duplicates(subset='사업자번호_norm', keep='first').copy()

print(f"\nHSM 그룹사 중복제거 후: {len(hsm_grp_dedup)}건")
print(f"HSM 계열사 중복제거 후: {len(hsm_sub_dedup)}건")

# === 4. Build HSM lookup by normalized 사업자번호 ===
# Use 그룹사 sheet as primary, supplement with 계열사 sheet
grp_lookup = hsm_grp_dedup.set_index('사업자번호_norm')
sub_lookup = hsm_sub_dedup.set_index('사업자번호_norm')

# === 5. Phase 1: Match on 사업자번호 ===
matched_bizno = []
unmatched = []

for idx, row in dj.iterrows():
    bizno = row['사업자번호_norm']
    if bizno and bizno in grp_lookup.index:
        matched_bizno.append((idx, bizno, 'grp', '사업자번호'))
    elif bizno and bizno in sub_lookup.index:
        matched_bizno.append((idx, bizno, 'sub', '사업자번호'))
    else:
        unmatched.append(idx)

print(f"\n=== Phase 1: 사업자번호 매칭 ===")
print(f"매칭 성공: {len(matched_bizno)}건")
print(f"미매칭: {len(unmatched)}건")

# === 6. Phase 2: Name normalization and matching for unmatched ===
def normalize_name(name):
    if pd.isna(name):
        return ''
    s = str(name).strip()
    # Remove common prefixes/suffixes
    s = re.sub(r'\(주\)|\(유\)|㈜|㈜|주식회사|유한회사|사단법인|재단법인|학교법인|의료법인|사회복지법인', '', s)
    s = re.sub(r'\(TB\)|\(tb\)', '', s)
    # Remove all whitespace and special chars
    s = re.sub(r'[\s\-\.\,\(\)\/·]', '', s)
    # Lowercase for English parts
    s = s.lower()
    return s.strip()

# Build name lookup from HSM
grp_name_lookup = {}
for bizno, row in grp_lookup.iterrows():
    name_norm = normalize_name(row.get('기업명', ''))
    if name_norm and name_norm not in grp_name_lookup:
        grp_name_lookup[name_norm] = bizno

sub_name_lookup = {}
for bizno, row in sub_lookup.iterrows():
    name_norm = normalize_name(row.get('계열사 기업명', ''))
    if name_norm and name_norm not in sub_name_lookup:
        sub_name_lookup[name_norm] = bizno

print(f"\nHSM 그룹사 정규화 기업명 수: {len(grp_name_lookup)}")
print(f"HSM 계열사 정규화 기업명 수: {len(sub_name_lookup)}")

# Phase 2a: Exact normalized name match
matched_name_exact = []
still_unmatched = []

for idx in unmatched:
    row = dj.loc[idx]
    name_norm = normalize_name(row['회사명'])
    if name_norm in grp_name_lookup:
        matched_name_exact.append((idx, grp_name_lookup[name_norm], 'grp', '기업명정규화'))
    elif name_norm in sub_name_lookup:
        matched_name_exact.append((idx, sub_name_lookup[name_norm], 'sub', '기업명정규화'))
    else:
        still_unmatched.append(idx)

print(f"\n=== Phase 2a: 정규화 기업명 매칭 ===")
print(f"매칭 성공: {len(matched_name_exact)}건")
print(f"미매칭: {len(still_unmatched)}건")

# Phase 2b: Fuzzy name matching for remaining
all_grp_names = list(grp_name_lookup.keys())
all_sub_names = list(sub_name_lookup.keys())
all_names = list(set(all_grp_names + all_sub_names))

matched_fuzzy = []
final_unmatched = []

for idx in still_unmatched:
    row = dj.loc[idx]
    name_norm = normalize_name(row['회사명'])
    if not name_norm:
        final_unmatched.append(idx)
        continue

    best_score = 0
    best_match = None
    best_source = None

    # Check against grp names
    for gname in all_grp_names:
        score = SequenceMatcher(None, name_norm, gname).ratio()
        if score > best_score:
            best_score = score
            best_match = gname
            best_source = 'grp'

    # Check against sub names
    for sname in all_sub_names:
        score = SequenceMatcher(None, name_norm, sname).ratio()
        if score > best_score:
            best_score = score
            best_match = sname
            best_source = 'sub'

    if best_score >= 0.75:
        if best_source == 'grp':
            matched_fuzzy.append((idx, grp_name_lookup[best_match], best_source, f'퍼지매칭({best_score:.2f})'))
        else:
            matched_fuzzy.append((idx, sub_name_lookup[best_match], best_source, f'퍼지매칭({best_score:.2f})'))
    else:
        final_unmatched.append(idx)

print(f"\n=== Phase 2b: 퍼지 기업명 매칭 (>=0.75) ===")
print(f"매칭 성공: {len(matched_fuzzy)}건")
print(f"최종 미매칭: {len(final_unmatched)}건")

# === 7. Combine all matches ===
all_matches = matched_bizno + matched_name_exact + matched_fuzzy

print(f"\n{'='*50}")
print(f"=== 최종 매핑 요약 ===")
print(f"전체: {len(dj)}건")
print(f"사업자번호 매칭: {len(matched_bizno)}건 ({len(matched_bizno)/len(dj)*100:.1f}%)")
print(f"기업명 정규화 매칭: {len(matched_name_exact)}건 ({len(matched_name_exact)/len(dj)*100:.1f}%)")
print(f"퍼지 매칭: {len(matched_fuzzy)}건 ({len(matched_fuzzy)/len(dj)*100:.1f}%)")
print(f"총 매칭: {len(all_matches)}건 ({len(all_matches)/len(dj)*100:.1f}%)")
print(f"미매칭: {len(final_unmatched)}건 ({len(final_unmatched)/len(dj)*100:.1f}%)")

# === 8. Build output DataFrame ===
result_rows = []

for match_idx, bizno_norm, source, method in all_matches:
    dj_row = dj.loc[match_idx]

    if source == 'grp':
        hsm_row = grp_lookup.loc[bizno_norm]
        grp_code = hsm_row.get('기업 코드', '')
        grp_name = hsm_row.get('기업명', '')
        sub_code = ''
        sub_name = ''
        biz_unit = hsm_row.get('영업단위', '')
        biz_unit_code = hsm_row.get('영업단위 코드', '')
        sales_rep_dept = hsm_row.get('영업대표부서', '')
        sales_rep_name = hsm_row.get('영업대표명', '')
        sales_rep_id = hsm_row.get('영업대표아이디', '')
        note = hsm_row.get('비고', '')
    else:
        hsm_row = sub_lookup.loc[bizno_norm]
        grp_code = hsm_row.get('그룹사 코드', '')
        grp_name = hsm_row.get('그룹사 기업명', '')
        sub_code = hsm_row.get('계열사 코드', '')
        sub_name = hsm_row.get('계열사 기업명', '')
        biz_unit = hsm_row.get('영업단위', '')
        biz_unit_code = hsm_row.get('영업단위 코드', '')
        sales_rep_dept = hsm_row.get('영업대표부서', '')
        sales_rep_name = hsm_row.get('영업대표명', '')
        sales_rep_id = hsm_row.get('영업대표아이디', '')
        note = hsm_row.get('비고', '')

    result_rows.append({
        'No': dj_row['No.'],
        '회사명(데일리잡)': dj_row['회사명'],
        '사업자번호': dj_row['사업자번호'],
        '주소지': dj_row['주소지'],
        '연락처': dj_row['연락처'],
        '직원수': dj_row['직원수'],
        '2020년 매출': dj_row.get('2020년 매출', ''),
        '2021년 매출': dj_row.get('2021년 매출', ''),
        '2022년 매출': dj_row.get('2022년 매출', ''),
        '2023년 매출': dj_row.get('2023년 매출', ''),
        '2024년 매출': dj_row.get('2024년 매출', ''),
        '2025년 매출': dj_row.get('2025년 매출', ''),
        '교육담당자 이름': dj_row.get('교육담당자 이름', ''),
        '교육담당자 직급': dj_row.get('교육담당자 직급', ''),
        '교육담당자 부서': dj_row.get('교육담당자 부서', ''),
        '교육담당자 전화': dj_row.get('교육담당자 전화', ''),
        '매칭방법': method,
        '매칭여부': 'Y',
        'HSM_그룹사코드': grp_code,
        'HSM_그룹사명': grp_name,
        'HSM_계열사코드': sub_code,
        'HSM_계열사명': sub_name,
        'HSM_영업단위': biz_unit,
        'HSM_영업단위코드': biz_unit_code,
        'HSM_영업대표부서': sales_rep_dept,
        'HSM_영업대표명': sales_rep_name,
        'HSM_영업대표아이디': sales_rep_id,
        'HSM_비고': note,
    })

# Add unmatched rows
for idx in final_unmatched:
    dj_row = dj.loc[idx]
    result_rows.append({
        'No': dj_row['No.'],
        '회사명(데일리잡)': dj_row['회사명'],
        '사업자번호': dj_row['사업자번호'],
        '주소지': dj_row['주소지'],
        '연락처': dj_row['연락처'],
        '직원수': dj_row['직원수'],
        '2020년 매출': dj_row.get('2020년 매출', ''),
        '2021년 매출': dj_row.get('2021년 매출', ''),
        '2022년 매출': dj_row.get('2022년 매출', ''),
        '2023년 매출': dj_row.get('2023년 매출', ''),
        '2024년 매출': dj_row.get('2024년 매출', ''),
        '2025년 매출': dj_row.get('2025년 매출', ''),
        '교육담당자 이름': dj_row.get('교육담당자 이름', ''),
        '교육담당자 직급': dj_row.get('교육담당자 직급', ''),
        '교육담당자 부서': dj_row.get('교육담당자 부서', ''),
        '교육담당자 전화': dj_row.get('교육담당자 전화', ''),
        '매칭방법': '',
        '매칭여부': 'N',
        'HSM_그룹사코드': '',
        'HSM_그룹사명': '',
        'HSM_계열사코드': '',
        'HSM_계열사명': '',
        'HSM_영업단위': '',
        'HSM_영업단위코드': '',
        'HSM_영업대표부서': '',
        'HSM_영업대표명': '',
        'HSM_영업대표아이디': '',
        'HSM_비고': '',
    })

result_df = pd.DataFrame(result_rows)

# === 9. Write to Excel with formatting ===
output_path = 'C:/Projects/legend-team/Data/데일리잡_HSM매핑결과.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Sheet 1: All data
    result_df.to_excel(writer, sheet_name='매핑결과(전체)', index=False)

    # Sheet 2: Matched only
    matched_df = result_df[result_df['매칭여부'] == 'Y']
    matched_df.to_excel(writer, sheet_name='매핑성공', index=False)

    # Sheet 3: Unmatched only
    unmatched_df = result_df[result_df['매칭여부'] == 'N']
    unmatched_df.to_excel(writer, sheet_name='미매칭', index=False)

    # Sheet 4: Summary
    summary_data = {
        '구분': ['전체 데일리잡', '사업자번호 매칭', '기업명 정규화 매칭', '퍼지 매칭(>=0.75)', '총 매칭', '미매칭'],
        '건수': [len(dj), len(matched_bizno), len(matched_name_exact), len(matched_fuzzy), len(all_matches), len(final_unmatched)],
        '비율': [
            '100.0%',
            f'{len(matched_bizno)/len(dj)*100:.1f}%',
            f'{len(matched_name_exact)/len(dj)*100:.1f}%',
            f'{len(matched_fuzzy)/len(dj)*100:.1f}%',
            f'{len(all_matches)/len(dj)*100:.1f}%',
            f'{len(final_unmatched)/len(dj)*100:.1f}%'
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='매핑요약', index=False)

# Apply formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(output_path)

header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
match_fill = PatternFill('solid', fgColor='E2EFDA')
nomatch_fill = PatternFill('solid', fgColor='FCE4EC')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    # Header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(size=10, name='맑은 고딕')
            cell.border = thin_border

    # Color code matched/unmatched rows in full sheet
    if ws_name == '매핑결과(전체)':
        match_col_idx = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == '매칭여부':
                match_col_idx = i
                break
        if match_col_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                if row[match_col_idx-1].value == 'Y':
                    for cell in row:
                        cell.fill = match_fill
                else:
                    for cell in row:
                        cell.fill = nomatch_fill

    # Auto column width (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 8), 40)

    # Freeze header
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

wb.save(output_path)
print(f"\n파일 저장 완료: {output_path}")

# Show some fuzzy match examples
if matched_fuzzy:
    print(f"\n=== 퍼지 매칭 샘플 (최대 20건) ===")
    for match_idx, bizno_norm, source, method in matched_fuzzy[:20]:
        dj_name = dj.loc[match_idx, '회사명']
        if source == 'grp' and bizno_norm in grp_lookup.index:
            hsm_name = grp_lookup.loc[bizno_norm, '기업명']
        elif bizno_norm in sub_lookup.index:
            hsm_name = sub_lookup.loc[bizno_norm, '계열사 기업명']
        else:
            hsm_name = '?'
        print(f"  {dj_name} → {hsm_name} [{method}]")

# Show some unmatched examples
if final_unmatched:
    print(f"\n=== 미매칭 샘플 (최대 20건) ===")
    for idx in final_unmatched[:20]:
        print(f"  {dj.loc[idx, '회사명']} (사업자번호: {dj.loc[idx, '사업자번호']})")
