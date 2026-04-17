import pandas as pd
import re
import sys
import io
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# === 1. Load data ===
print("데이터 로딩 중...")
dj = pd.read_excel('C:/Projects/legend-team/Data/20260402_데일리잡 리스트_검토완료.xlsx', dtype=str)
hsm_grp = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='그룹사', dtype=str)
hsm_sub = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='계열사', dtype=str)

print(f"데일리잡: {len(dj)}건")
print(f"HSM 그룹사: {len(hsm_grp)}건, HSM 계열사: {len(hsm_sub)}건")

# === 2. Normalize functions ===
def norm_bizno(val):
    if pd.isna(val): return ''
    return re.sub(r'[^0-9]', '', str(val).strip())

def norm_name(name):
    if pd.isna(name): return ''
    s = str(name).strip()
    s = re.sub(r'\(주\)|\(유\)|㈜|주식회사|유한회사|사단법인|재단법인|학교법인|의료법인|사회복지법인|농업회사법인|영농조합법인', '', s)
    s = re.sub(r'\(TB\)|\(tb\)', '', s)
    s = re.sub(r'[\s\-\.\,\(\)\/·\']', '', s)
    s = s.lower().strip()
    return s

def make_bigrams(s):
    if len(s) < 2: return set([s]) if s else set()
    return set(s[i:i+2] for i in range(len(s)-1))

def bigram_similarity(s1, s2):
    if not s1 or not s2: return 0.0
    b1, b2 = make_bigrams(s1), make_bigrams(s2)
    if not b1 or not b2: return 0.0
    return 2.0 * len(b1 & b2) / (len(b1) + len(b2))

# === 3. Normalize all 사업자번호 ===
dj['_bizno'] = dj['사업자번호'].apply(norm_bizno)
hsm_grp['_bizno'] = hsm_grp['사업자번호'].apply(norm_bizno)
hsm_sub['_bizno'] = hsm_sub['사업자번호'].apply(norm_bizno)

# === 4. Deduplicate HSM ===
grp_dedup = hsm_grp.drop_duplicates(subset='_bizno', keep='first').set_index('_bizno')
sub_dedup = hsm_sub.drop_duplicates(subset='_bizno', keep='first').set_index('_bizno')
print(f"HSM 그룹사 고유: {len(grp_dedup)}건, HSM 계열사 고유: {len(sub_dedup)}건")

# === 5. Build combined HSM name lookup (normalized name → bizno + source) ===
hsm_names = {}  # norm_name -> (bizno, source)
for bizno, row in grp_dedup.iterrows():
    nn = norm_name(row.get('기업명', ''))
    if nn and nn not in hsm_names:
        hsm_names[nn] = (bizno, 'grp')
for bizno, row in sub_dedup.iterrows():
    nn = norm_name(row.get('계열사 기업명', ''))
    if nn and nn not in hsm_names:
        hsm_names[nn] = (bizno, 'sub')

print(f"HSM 정규화 기업명 수: {len(hsm_names)}")

# Build bigram index for fast fuzzy lookup
print("바이그램 인덱스 구축 중...")
bigram_index = defaultdict(list)  # bigram -> list of norm_names
for nn in hsm_names:
    for bg in make_bigrams(nn):
        bigram_index[bg].append(nn)

# === 6. Matching ===
matches = []  # (dj_idx, bizno, source, method)
unmatched_indices = []

print("매칭 시작...")
bizno_match_count = 0
name_exact_count = 0
fuzzy_match_count = 0

for idx, row in dj.iterrows():
    bizno = row['_bizno']
    name = norm_name(row['회사명'])

    # Phase 1: 사업자번호
    if bizno and bizno in grp_dedup.index:
        matches.append((idx, bizno, 'grp', '사업자번호'))
        bizno_match_count += 1
        continue
    if bizno and bizno in sub_dedup.index:
        matches.append((idx, bizno, 'sub', '사업자번호'))
        bizno_match_count += 1
        continue

    # Phase 2a: exact normalized name
    if name and name in hsm_names:
        b, s = hsm_names[name]
        matches.append((idx, b, s, '기업명정규화'))
        name_exact_count += 1
        continue

    # Phase 2b: fuzzy via bigram pre-filter
    if name:
        query_bigrams = make_bigrams(name)
        candidate_counts = defaultdict(int)
        for bg in query_bigrams:
            for cand in bigram_index.get(bg, []):
                candidate_counts[cand] += 1

        # Get top candidates (at least 2 shared bigrams)
        top_cands = sorted(candidate_counts.items(), key=lambda x: -x[1])[:50]

        best_score = 0
        best_cand = None
        for cand_name, _ in top_cands:
            score = bigram_similarity(name, cand_name)
            if score > best_score:
                best_score = score
                best_cand = cand_name

        if best_score >= 0.6 and best_cand:
            b, s = hsm_names[best_cand]
            matches.append((idx, b, s, f'퍼지({best_score:.2f})'))
            fuzzy_match_count += 1
            continue

    unmatched_indices.append(idx)

    if (idx + 1) % 500 == 0:
        print(f"  진행: {idx+1}/{len(dj)}...")

print(f"\n{'='*50}")
print(f"=== 매핑 결과 요약 ===")
print(f"전체: {len(dj)}건")
print(f"사업자번호 매칭: {bizno_match_count}건 ({bizno_match_count/len(dj)*100:.1f}%)")
print(f"기업명정규화 매칭: {name_exact_count}건 ({name_exact_count/len(dj)*100:.1f}%)")
print(f"퍼지 매칭: {fuzzy_match_count}건 ({fuzzy_match_count/len(dj)*100:.1f}%)")
total_matched = bizno_match_count + name_exact_count + fuzzy_match_count
print(f"총 매칭: {total_matched}건 ({total_matched/len(dj)*100:.1f}%)")
print(f"미매칭: {len(unmatched_indices)}건 ({len(unmatched_indices)/len(dj)*100:.1f}%)")

# === 7. Build output ===
print("\n엑셀 파일 생성 중...")

def get_hsm_info(bizno, source):
    if source == 'grp' and bizno in grp_dedup.index:
        r = grp_dedup.loc[bizno]
        return {
            'HSM_그룹사코드': r.get('기업 코드', ''),
            'HSM_그룹사명': r.get('기업명', ''),
            'HSM_계열사코드': '', 'HSM_계열사명': '',
            'HSM_영업단위': r.get('영업단위', ''),
            'HSM_영업단위코드': r.get('영업단위 코드', ''),
            'HSM_영업대표부서': r.get('영업대표부서', ''),
            'HSM_영업대표명': r.get('영업대표명', ''),
            'HSM_영업대표아이디': r.get('영업대표아이디', ''),
            'HSM_비고': r.get('비고', ''),
        }
    elif source == 'sub' and bizno in sub_dedup.index:
        r = sub_dedup.loc[bizno]
        return {
            'HSM_그룹사코드': r.get('그룹사 코드', ''),
            'HSM_그룹사명': r.get('그룹사 기업명', ''),
            'HSM_계열사코드': r.get('계열사 코드', ''),
            'HSM_계열사명': r.get('계열사 기업명', ''),
            'HSM_영업단위': r.get('영업단위', ''),
            'HSM_영업단위코드': r.get('영업단위 코드', ''),
            'HSM_영업대표부서': r.get('영업대표부서', ''),
            'HSM_영업대표명': r.get('영업대표명', ''),
            'HSM_영업대표아이디': r.get('영업대표아이디', ''),
            'HSM_비고': r.get('비고', ''),
        }
    return {k: '' for k in ['HSM_그룹사코드','HSM_그룹사명','HSM_계열사코드','HSM_계열사명',
                             'HSM_영업단위','HSM_영업단위코드','HSM_영업대표부서','HSM_영업대표명','HSM_영업대표아이디','HSM_비고']}

rows = []
dj_cols = ['No.','회사명','사업자번호','주소지','연락처','직원수',
           '2020년 매출','2021년 매출','2022년 매출','2023년 매출','2024년 매출','2025년 매출',
           '교육담당자 이름','교육담당자 직급','교육담당자 부서','교육담당자 전화']

for dj_idx, bizno, source, method in matches:
    r = dj.loc[dj_idx]
    row_data = {c: r.get(c, '') for c in dj_cols}
    row_data['매칭여부'] = 'Y'
    row_data['매칭방법'] = method
    row_data.update(get_hsm_info(bizno, source))
    rows.append(row_data)

for dj_idx in unmatched_indices:
    r = dj.loc[dj_idx]
    row_data = {c: r.get(c, '') for c in dj_cols}
    row_data['매칭여부'] = 'N'
    row_data['매칭방법'] = ''
    row_data.update({k: '' for k in ['HSM_그룹사코드','HSM_그룹사명','HSM_계열사코드','HSM_계열사명',
                                      'HSM_영업단위','HSM_영업단위코드','HSM_영업대표부서','HSM_영업대표명','HSM_영업대표아이디','HSM_비고']})
    rows.append(row_data)

result = pd.DataFrame(rows)

output_path = 'C:/Projects/legend-team/Data/데일리잡_HSM매핑결과.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    result.to_excel(writer, sheet_name='매핑결과(전체)', index=False)
    result[result['매칭여부']=='Y'].to_excel(writer, sheet_name='매핑성공', index=False)
    result[result['매칭여부']=='N'].to_excel(writer, sheet_name='미매칭', index=False)

    summary = pd.DataFrame({
        '구분': ['전체 데일리잡', '사업자번호 매칭', '기업명정규화 매칭', '퍼지 매칭', '총 매칭', '미매칭'],
        '건수': [len(dj), bizno_match_count, name_exact_count, fuzzy_match_count, total_matched, len(unmatched_indices)],
        '비율': ['100.0%', f'{bizno_match_count/len(dj)*100:.1f}%', f'{name_exact_count/len(dj)*100:.1f}%',
                f'{fuzzy_match_count/len(dj)*100:.1f}%', f'{total_matched/len(dj)*100:.1f}%', f'{len(unmatched_indices)/len(dj)*100:.1f}%']
    })
    summary.to_excel(writer, sheet_name='매핑요약', index=False)

# Formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(output_path)
hdr_fill = PatternFill('solid', fgColor='4472C4')
hdr_font = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
y_fill = PatternFill('solid', fgColor='E2EFDA')
n_fill = PatternFill('solid', fgColor='FCE4EC')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
body_font = Font(size=10, name='맑은 고딕')

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font
            cell.border = thin

    if ws.title == '매핑결과(전체)':
        mi = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == '매칭여부':
                mi = i; break
        if mi:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                fill = y_fill if row[mi-1].value == 'Y' else n_fill
                for cell in row:
                    cell.fill = fill

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

wb.save(output_path)

print(f"\n파일 저장 완료: {output_path}")

# Show fuzzy match samples
fuzzy_matches = [(i, b, s, m) for i, b, s, m in matches if '퍼지' in m]
if fuzzy_matches:
    print(f"\n=== 퍼지 매칭 샘플 (최대 20건) ===")
    for dj_idx, bizno, source, method in fuzzy_matches[:20]:
        dj_name = dj.loc[dj_idx, '회사명']
        if source == 'grp' and bizno in grp_dedup.index:
            hsm_name = grp_dedup.loc[bizno, '기업명']
        elif bizno in sub_dedup.index:
            hsm_name = sub_dedup.loc[bizno, '계열사 기업명']
        else:
            hsm_name = '?'
        print(f"  {dj_name} → {hsm_name} [{method}]")

if unmatched_indices:
    print(f"\n=== 미매칭 샘플 (최대 20건) ===")
    for idx in unmatched_indices[:20]:
        print(f"  {dj.loc[idx, '회사명']} (사업자번호: {dj.loc[idx, '사업자번호']})")
