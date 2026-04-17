import pandas as pd
import re
import sys
import io
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("데이터 로딩 중...")
dj = pd.read_excel('C:/Projects/legend-team/Data/20260402_데일리잡 리스트_검토완료.xlsx', dtype=str)
hsm_grp = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='그룹사', dtype=str)
hsm_sub = pd.read_excel('C:/Projects/legend-team/Data/HSM 계열사_그룹사 관련 데이터_원본.xlsx', sheet_name='계열사', dtype=str)

def norm_bizno(val):
    if pd.isna(val): return ''
    return re.sub(r'[^0-9]', '', str(val).strip())

def norm_name(name):
    if pd.isna(name): return ''
    s = str(name).strip()
    s = re.sub(r'\(주\)|\(유\)|㈜|주식회사|유한회사|사단법인|재단법인|학교법인|의료법인|사회복지법인|농업회사법인|영농조합법인|합자회사|합명회사', '', s)
    s = re.sub(r'\(TB\)|\(tb\)', '', s)
    s = re.sub(r'[\s\-\.\,\(\)\/·\'\"]', '', s)
    s = s.lower().strip()
    return s

def make_bigrams(s):
    if len(s) < 2: return set([s]) if s else set()
    return set(s[i:i+2] for i in range(len(s)-1))

def bigram_sim(s1, s2):
    if not s1 or not s2: return 0.0
    b1, b2 = make_bigrams(s1), make_bigrams(s2)
    if not b1 or not b2: return 0.0
    return 2.0 * len(b1 & b2) / (len(b1) + len(b2))

# Normalize
dj['_biz'] = dj['사업자번호'].apply(norm_bizno)
hsm_grp['_biz'] = hsm_grp['사업자번호'].apply(norm_bizno)
hsm_sub['_biz'] = hsm_sub['사업자번호'].apply(norm_bizno)

# Dedup HSM
grp = hsm_grp.drop_duplicates(subset='_biz', keep='first').set_index('_biz')
sub = hsm_sub.drop_duplicates(subset='_biz', keep='first').set_index('_biz')

# Name lookups
hsm_names = {}
for biz, row in grp.iterrows():
    nn = norm_name(row.get('기업명', ''))
    if nn and nn not in hsm_names:
        hsm_names[nn] = (biz, 'grp', row.get('기업명', ''))
for biz, row in sub.iterrows():
    nn = norm_name(row.get('계열사 기업명', ''))
    if nn and nn not in hsm_names:
        hsm_names[nn] = (biz, 'sub', row.get('계열사 기업명', ''))

# Bigram index
bg_idx = defaultdict(list)
for nn in hsm_names:
    for bg in make_bigrams(nn):
        bg_idx[bg].append(nn)

print(f"데일리잡: {len(dj)}건, HSM 기업(그룹사): {len(grp)}건, HSM 기업(계열사): {len(sub)}건")

# Matching
matches = []  # (dj_idx, biz, source, method, confidence, hsm_name_matched)

biz_cnt = name_cnt = fuzzy_high_cnt = fuzzy_mid_cnt = fuzzy_low_cnt = 0
unmatched = []

for idx, row in dj.iterrows():
    biz = row['_biz']
    nm = norm_name(row['회사명'])

    # Phase 1: 사업자번호
    if biz and biz in grp.index:
        matches.append((idx, biz, 'grp', '사업자번호', '확정', grp.loc[biz].get('기업명','')))
        biz_cnt += 1; continue
    if biz and biz in sub.index:
        matches.append((idx, biz, 'sub', '사업자번호', '확정', sub.loc[biz].get('계열사 기업명','')))
        biz_cnt += 1; continue

    # Phase 2: exact normalized name
    if nm and nm in hsm_names:
        b, s, orig = hsm_names[nm]
        matches.append((idx, b, s, '기업명정규화', '확정', orig))
        name_cnt += 1; continue

    # Phase 3: fuzzy bigram
    if nm:
        qbg = make_bigrams(nm)
        cands = defaultdict(int)
        for bg in qbg:
            for c in bg_idx.get(bg, []):
                cands[c] += 1
        top = sorted(cands.items(), key=lambda x: -x[1])[:30]

        best_score = 0; best_cand = None
        for cn, _ in top:
            sc = bigram_sim(nm, cn)
            if sc > best_score:
                best_score = sc; best_cand = cn

        if best_score >= 0.85 and best_cand:
            b, s, orig = hsm_names[best_cand]
            matches.append((idx, b, s, f'퍼지({best_score:.2f})', '높음', orig))
            fuzzy_high_cnt += 1; continue
        elif best_score >= 0.70 and best_cand:
            b, s, orig = hsm_names[best_cand]
            matches.append((idx, b, s, f'퍼지({best_score:.2f})', '중간(검수필요)', orig))
            fuzzy_mid_cnt += 1; continue
        elif best_score >= 0.55 and best_cand:
            b, s, orig = hsm_names[best_cand]
            matches.append((idx, b, s, f'퍼지({best_score:.2f})', '낮음(검수필요)', orig))
            fuzzy_low_cnt += 1; continue

    unmatched.append(idx)

total_m = biz_cnt + name_cnt + fuzzy_high_cnt + fuzzy_mid_cnt + fuzzy_low_cnt
confident = biz_cnt + name_cnt + fuzzy_high_cnt

print(f"\n{'='*60}")
print(f"=== 매핑 결과 요약 ===")
print(f"전체: {len(dj)}건")
print(f"")
print(f"[확정 매칭]")
print(f"  사업자번호 매칭:    {biz_cnt:>5}건 ({biz_cnt/len(dj)*100:5.1f}%)")
print(f"  기업명정규화 매칭:  {name_cnt:>5}건 ({name_cnt/len(dj)*100:5.1f}%)")
print(f"  퍼지(>=0.85):       {fuzzy_high_cnt:>5}건 ({fuzzy_high_cnt/len(dj)*100:5.1f}%)")
print(f"  → 확정 소계:        {confident:>5}건 ({confident/len(dj)*100:5.1f}%)")
print(f"")
print(f"[검수 필요]")
print(f"  퍼지(0.70~0.84):    {fuzzy_mid_cnt:>5}건 ({fuzzy_mid_cnt/len(dj)*100:5.1f}%)")
print(f"  퍼지(0.55~0.69):    {fuzzy_low_cnt:>5}건 ({fuzzy_low_cnt/len(dj)*100:5.1f}%)")
print(f"  → 검수 소계:        {fuzzy_mid_cnt+fuzzy_low_cnt:>5}건 ({(fuzzy_mid_cnt+fuzzy_low_cnt)/len(dj)*100:5.1f}%)")
print(f"")
print(f"[미매칭]              {len(unmatched):>5}건 ({len(unmatched)/len(dj)*100:5.1f}%)")
print(f"")
print(f"총 매칭(검수 포함):   {total_m:>5}건 ({total_m/len(dj)*100:5.1f}%)")

# Build output
print("\n엑셀 파일 생성 중...")

def hsm_info(biz, src):
    empty = {k: '' for k in ['HSM_그룹사코드','HSM_그룹사명','HSM_계열사코드','HSM_계열사명',
                              'HSM_영업단위','HSM_영업단위코드','HSM_영업대표부서','HSM_영업대표명','HSM_영업대표아이디','HSM_비고']}
    if src == 'grp' and biz in grp.index:
        r = grp.loc[biz]
        return {'HSM_그룹사코드': r.get('기업 코드',''), 'HSM_그룹사명': r.get('기업명',''),
                'HSM_계열사코드': '', 'HSM_계열사명': '',
                'HSM_영업단위': r.get('영업단위',''), 'HSM_영업단위코드': r.get('영업단위 코드',''),
                'HSM_영업대표부서': r.get('영업대표부서',''), 'HSM_영업대표명': r.get('영업대표명',''),
                'HSM_영업대표아이디': r.get('영업대표아이디',''), 'HSM_비고': r.get('비고','')}
    elif src == 'sub' and biz in sub.index:
        r = sub.loc[biz]
        return {'HSM_그룹사코드': r.get('그룹사 코드',''), 'HSM_그룹사명': r.get('그룹사 기업명',''),
                'HSM_계열사코드': r.get('계열사 코드',''), 'HSM_계열사명': r.get('계열사 기업명',''),
                'HSM_영업단위': r.get('영업단위',''), 'HSM_영업단위코드': r.get('영업단위 코드',''),
                'HSM_영업대표부서': r.get('영업대표부서',''), 'HSM_영업대표명': r.get('영업대표명',''),
                'HSM_영업대표아이디': r.get('영업대표아이디',''), 'HSM_비고': r.get('비고','')}
    return empty

dj_cols = ['No.','회사명','사업자번호','주소지','연락처','직원수',
           '2020년 매출','2021년 매출','2022년 매출','2023년 매출','2024년 매출','2025년 매출',
           '교육담당자 이름','교육담당자 직급','교육담당자 부서','교육담당자 전화']

rows = []
for di, biz, src, meth, conf, hsm_nm in matches:
    r = dj.loc[di]
    d = {c: r.get(c,'') for c in dj_cols}
    d['매칭여부'] = 'Y'
    d['매칭방법'] = meth
    d['신뢰도'] = conf
    d['HSM_매칭기업명'] = hsm_nm
    d.update(hsm_info(biz, src))
    rows.append(d)

for di in unmatched:
    r = dj.loc[di]
    d = {c: r.get(c,'') for c in dj_cols}
    d['매칭여부'] = 'N'
    d['매칭방법'] = ''
    d['신뢰도'] = ''
    d['HSM_매칭기업명'] = ''
    d.update({k:'' for k in ['HSM_그룹사코드','HSM_그룹사명','HSM_계열사코드','HSM_계열사명',
                              'HSM_영업단위','HSM_영업단위코드','HSM_영업대표부서','HSM_영업대표명','HSM_영업대표아이디','HSM_비고']})
    rows.append(d)

result = pd.DataFrame(rows)
output = 'C:/Projects/legend-team/Data/데일리잡_HSM매핑결과.xlsx'

with pd.ExcelWriter(output, engine='openpyxl') as w:
    result.to_excel(w, sheet_name='매핑결과(전체)', index=False)
    result[result['매칭여부']=='Y'].to_excel(w, sheet_name='매핑성공', index=False)
    result[result['신뢰도'].isin(['중간(검수필요)','낮음(검수필요)'])].to_excel(w, sheet_name='검수필요', index=False)
    result[result['매칭여부']=='N'].to_excel(w, sheet_name='미매칭', index=False)

    pd.DataFrame({
        '구분': ['전체 데일리잡', '', '[확정 매칭]', '  사업자번호', '  기업명정규화', '  퍼지(>=0.85)', '  확정 소계', '',
                '[검수 필요]', '  퍼지(0.70~0.84)', '  퍼지(0.55~0.69)', '  검수 소계', '', '[미매칭]', '', '총 매칭(검수포함)'],
        '건수': [len(dj), '', '', biz_cnt, name_cnt, fuzzy_high_cnt, confident, '',
                '', fuzzy_mid_cnt, fuzzy_low_cnt, fuzzy_mid_cnt+fuzzy_low_cnt, '', len(unmatched), '', total_m],
        '비율': ['100.0%', '', '', f'{biz_cnt/len(dj)*100:.1f}%', f'{name_cnt/len(dj)*100:.1f}%',
                f'{fuzzy_high_cnt/len(dj)*100:.1f}%', f'{confident/len(dj)*100:.1f}%', '',
                '', f'{fuzzy_mid_cnt/len(dj)*100:.1f}%', f'{fuzzy_low_cnt/len(dj)*100:.1f}%',
                f'{(fuzzy_mid_cnt+fuzzy_low_cnt)/len(dj)*100:.1f}%', '', f'{len(unmatched)/len(dj)*100:.1f}%', '',
                f'{total_m/len(dj)*100:.1f}%']
    }).to_excel(w, sheet_name='매핑요약', index=False)

# Formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(output)
hf = PatternFill('solid', fgColor='4472C4')
hfont = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
green = PatternFill('solid', fgColor='E2EFDA')
yellow = PatternFill('solid', fgColor='FFF2CC')
red = PatternFill('solid', fgColor='FCE4EC')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
bf = Font(size=10, name='맑은 고딕')

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font = hfont; cell.fill = hf
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = bf; cell.border = thin

    if ws.title in ('매핑결과(전체)', '매핑성공'):
        ci = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == '신뢰도': ci = i; break
        mi = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == '매칭여부': mi = i; break
        if ci and mi:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                mv = row[mi-1].value
                cv = row[ci-1].value or ''
                if mv == 'N':
                    fill = red
                elif '검수' in cv:
                    fill = yellow
                else:
                    fill = green
                for cell in row:
                    cell.fill = fill

    for col in ws.columns:
        lt = col[0].column_letter
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[lt].width = min(max(mx + 2, 8), 40)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

wb.save(output)
print(f"\n파일 저장: {output}")

# Samples
fuzzy_samples = [(di,b,s,m,c,h) for di,b,s,m,c,h in matches if '퍼지' in m]
for conf_label in ['높음', '중간(검수필요)', '낮음(검수필요)']:
    samples = [(di,m,c,h) for di,b,s,m,c,h in fuzzy_samples if c == conf_label][:10]
    if samples:
        print(f"\n=== 퍼지 매칭 샘플 [{conf_label}] ===")
        for di, m, c, h in samples:
            print(f"  {dj.loc[di,'회사명']} → {h} [{m}]")

if unmatched:
    print(f"\n=== 미매칭 샘플 (최대 15건) ===")
    for di in unmatched[:15]:
        print(f"  {dj.loc[di,'회사명']} ({dj.loc[di,'사업자번호']})")
